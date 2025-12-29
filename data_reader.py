import openpyxl 
import csv
import os

class ExcelDataReader: 
    def __init__(self, filepath): 
        """ 
        初始化时，加载 Excel 工作簿。 
        """ 
        self.workbook = None
        self.filepath = filepath
        self.file_type = None
        
        # 检查文件类型
        if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
            self.file_type = 'excel'
            try:
                self.workbook = openpyxl.load_workbook(filepath, data_only=True)
                print(f"成功加载 Excel: {filepath}")
            except FileNotFoundError:
                print(f"错误：找不到文件 {filepath}")
            except Exception as e:
                print(f"加载 Excel 出错: {e}")
        elif filepath.endswith('.csv'):
            self.file_type = 'csv'
            print(f"识别到 CSV 文件: {filepath}")
        else:
            print(f"错误：不支持的文件类型 {filepath}")

    def find_value_by_label(self, sheet_name, label_name, column=None, search_direction='right',
                           exact_match=False, case_sensitive=False, max_rows=None):
        """
        通用函数：遍历 Excel 的某一列或整个工作表，找到包含 label_name 的单元格，然后返回其相邻单元格的值。
        摒弃硬坐标定位，使代码更健壮，能应对 Excel 格式的微调。

        Args:
            sheet_name: 工作表名称
            label_name: 要查找的标签文本
            column: 指定列（如'A', 'B'），如果为None则搜索整个工作表
            search_direction: 搜索方向 - 'right'（右侧）、'left'（左侧）、'below'（下方）、'above'（上方）
            exact_match: 是否要求精确匹配（False表示包含匹配）
            case_sensitive: 是否区分大小写
            max_rows: 最大搜索行数，None表示搜索所有行

        Returns:
            找到的值，如果没找到返回None
        """
        if not self.workbook or self.file_type != 'excel':
            return None

        try:
            sheet = self.workbook[sheet_name]
            if sheet_name not in self.workbook.sheetnames:
                print(f"错误：找不到工作表 {sheet_name}")
                return None

            # 准备标签文本
            search_text = str(label_name) if case_sensitive else str(label_name).lower()

            # 搜索范围
            search_range = []
            if column:
                # 搜索指定列
                column_index = openpyxl.utils.column_index_from_string(column)
                max_search_rows = min(max_rows or sheet.max_row, sheet.max_row)
                for row in range(1, max_search_rows + 1):
                    search_range.append(sheet.cell(row=row, column=column_index))
            else:
                # 搜索整个工作表
                max_search_rows = min(max_rows or sheet.max_row, sheet.max_row)
                for row in range(1, max_search_rows + 1):
                    for col in range(1, sheet.max_column + 1):
                        search_range.append(sheet.cell(row=row, column=col))

            # 搜索匹配的单元格
            matched_cells = []
            for cell in search_range:
                if cell.value is not None:
                    cell_text = str(cell.value)
                    compare_text = cell_text if case_sensitive else cell_text.lower()

                    # 匹配逻辑
                    if exact_match:
                        if search_text == compare_text:
                            matched_cells.append(cell)
                    else:
                        if search_text in compare_text:
                            matched_cells.append(cell)

            if not matched_cells:
                print(f"警告：在 {sheet_name} 中未找到包含 '{label_name}' 的单元格")
                return None

            # 返回第一个匹配单元格相邻的值
            target_cell = matched_cells[0]
            value_cell = None

            if search_direction == 'right':
                value_cell = sheet.cell(row=target_cell.row, column=target_cell.column + 1)
            elif search_direction == 'left':
                if target_cell.column > 1:
                    value_cell = sheet.cell(row=target_cell.row, column=target_cell.column - 1)
            elif search_direction == 'below':
                value_cell = sheet.cell(row=target_cell.row + 1, column=target_cell.column)
            elif search_direction == 'above':
                if target_cell.row > 1:
                    value_cell = sheet.cell(row=target_cell.row - 1, column=target_cell.column)

            return value_cell.value if value_cell and value_cell.value is not None else None

        except Exception as e:
            print(f"查找标签 '{label_name}' 时出错: {e}")
            return None

    def _find_value_next_to(self, sheet_name, keyword): 
        """ 
        私有方法，用于实现向后兼容。
        在指定的工作表中查找一个关键词，并返回其右侧单元格的值。
        """ 
        if not self.workbook: 
            return None 
            
        try: 
            sheet = self.workbook[sheet_name] 
            for row in sheet.iter_rows(): 
                for cell in row: 
                    if cell.value == keyword: 
                        # 找到了关键词！返回它右边一列的值 
                        value_cell = sheet.cell(row=cell.row, column=cell.column + 1) 
                        return value_cell.value 
            print(f"警告：在 {sheet_name} 中未找到关键词 '{keyword}'") 
            return None 
        except KeyError: 
            print(f"错误：找不到工作表 {sheet_name}") 
            return None 
        except Exception as e: 
            print(f"查找关键词 '{keyword}' 时出错: {e}") 
            return None
            
    def _find_value_below(self, sheet_name, keyword):
        """ 
        在指定的工作表中查找关键词，并返回其下方单元格的值。
        """
        if not self.workbook: 
            return None 
            
        try: 
            sheet = self.workbook[sheet_name] 
            for row in sheet.iter_rows(): 
                for cell in row: 
                    if cell.value == keyword: 
                        # 找到了关键词！返回它下方单元格的值 
                        value_cell = sheet.cell(row=cell.row + 1, column=cell.column) 
                        return value_cell.value 
            print(f"警告：在 {sheet_name} 中未找到关键词 '{keyword}'") 
            return None 
        except Exception as e: 
            print(f"查找关键词 '{keyword}' 下方值时出错: {e}") 
            return None
            
    def _find_value_by_content(self, sheet_name, keyword_substring):
        """ 
        在指定的工作表中查找包含关键词子串的单元格，并返回其下方单元格的值。
        用于模糊匹配，如'范围三'可能出现在不同格式的单元格中。
        """
        if not self.workbook: 
            return None 
            
        try: 
            sheet = self.workbook[sheet_name] 
            for row in sheet.iter_rows(): 
                for cell in row: 
                    if cell.value is not None and keyword_substring in str(cell.value): 
                        # 找到了包含关键词的单元格！返回它下方单元格的值 
                        value_cell = sheet.cell(row=cell.row + 1, column=cell.column) 
                        return value_cell.value 
            print(f"警告：在 {sheet_name} 中未找到包含 '{keyword_substring}' 的单元格") 
            return None 
        except Exception as e: 
            print(f"查找包含关键词 '{keyword_substring}' 的单元格时出错: {e}") 
            return None 

    def read_to_list_of_dicts(self, sheet_name=None, header_row=1, start_row=None,
                             end_row=None, skip_empty_rows=True, clean_headers=True):
        """
        将 CSV/Excel 文件中的数据转换为列表字典格式。
        能够处理各种数据格式，支持灵活的表头和数据行配置。

        Args:
            sheet_name: Excel工作表名称（CSV文件不需要）
            header_row: 表头所在行（默认第1行）
            start_row: 数据开始行（默认header_row+1）
            end_row: 数据结束行（默认工作表末尾）
            skip_empty_rows: 是否跳过空行
            clean_headers: 是否清理表头（去空格、标准化）

        Returns:
            列表字典格式: [{"列名1": 值1, "列名2": 值2, ...}, ...]
        """
        result = []

        if self.file_type == 'csv':
            # 处理 CSV 文件
            try:
                # 尝试多种编码
                encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin-1']
                csv_data = None

                for encoding in encodings:
                    try:
                        with open(self.filepath, 'r', encoding=encoding, newline='') as csvfile:
                            # 读取所有行
                            lines = csvfile.readlines()

                            # 处理表头行
                            if header_row <= len(lines):
                                header_line = lines[header_row - 1].strip()
                                headers = [h.strip() if clean_headers else h for h in header_line.split(',')]

                                # 确保表头不为空
                                for i, h in enumerate(headers):
                                    if not h or h.strip() == '':
                                        headers[i] = f"column_{i+1}"

                                # 处理数据行
                                data_start = start_row or header_row + 1
                                data_end = min(end_row or len(lines), len(lines))

                                for line_num in range(data_start, data_end + 1):
                                    if line_num <= len(lines):
                                        line = lines[line_num - 1].strip()
                                        if line:  # 非空行
                                            values = [v.strip() for v in line.split(',')]

                                            # 创建行字典
                                            row_dict = {}
                                            for i, header in enumerate(headers):
                                                value = values[i] if i < len(values) else None
                                                row_dict[header] = self._clean_cell_value(value)

                                            # 检查是否跳过空行
                                            if not skip_empty_rows or any(v is not None and v != '' for v in row_dict.values()):
                                                result.append(row_dict)
                                csv_data = True
                                break
                    except UnicodeDecodeError:
                        continue

                if csv_data is None:
                    print(f"无法使用任何编码读取 CSV 文件: {self.filepath}")
                else:
                    print(f"成功从 CSV 文件读取 {len(result)} 行数据")

            except Exception as e:
                print(f"读取 CSV 文件时出错: {e}")

        elif self.file_type == 'excel' and sheet_name:
            # 处理 Excel 文件
            if not self.workbook:
                return result

            try:
                if sheet_name not in self.workbook.sheetnames:
                    print(f"错误：找不到工作表 {sheet_name}")
                    return result

                sheet = self.workbook[sheet_name]

                # 获取表头
                headers = []
                header_row_obj = sheet[header_row]
                for cell in header_row_obj:
                    if cell.value is not None:
                        header_text = str(cell.value).strip()
                        if clean_headers:
                            # 清理表头：去空格、标准化
                            header_text = header_text.replace(' ', '_').replace('\n', '_').strip()
                        headers.append(header_text if header_text else f"column_{cell.column}")
                    else:
                        headers.append(f"column_{cell.column}")

                # 确保表头不为空
                for i, h in enumerate(headers):
                    if not h or h.strip() == '':
                        headers[i] = f"column_{i+1}"

                # 读取数据行
                data_start = start_row or header_row + 1
                data_end = min(end_row or sheet.max_row, sheet.max_row)

                for row in range(data_start, data_end + 1):
                    row_dict = {}
                    has_data = False

                    for col in range(1, len(headers) + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        cleaned_value = self._clean_cell_value(cell_value)
                        row_dict[headers[col-1]] = cleaned_value

                        if cleaned_value is not None and cleaned_value != '':
                            has_data = True

                    # 根据参数决定是否跳过空行
                    if not skip_empty_rows or has_data:
                        result.append(row_dict)

                print(f"成功从 Excel 工作表 {sheet_name} 读取 {len(result)} 行数据")

            except Exception as e:
                print(f"读取 Excel 工作表 {sheet_name} 时出错: {e}")

        else:
            print(f"错误：文件类型 {self.file_type} 或缺少必要参数")

        return result

    def _clean_cell_value(self, value):
        """
        清理和标准化单元格值

        Args:
            value: 原始单元格值

        Returns:
            清理后的值
        """
        if value is None:
            return None

        # 处理数字
        if isinstance(value, (int, float)):
            return value

        # 处理字符串
        if isinstance(value, str):
            # 去除首尾空格和换行符
            cleaned = value.strip()

            # 空字符串转为None
            if cleaned == '':
                return None

            # 尝试转换为数字
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return cleaned

        return value

    def read_emission_data_csv(self, csv_path='减排行动统计.csv'):
        """
        读取减排行动统计CSV文件，提取所有模板需要的变量

        Args:
            csv_path: CSV文件路径

        Returns:
            包含所有32个模板变量的字典
        """
        import os

        # 如果传入的filepath是CSV文件，使用它
        if self.file_type == 'csv' and self.filepath:
            csv_path = self.filepath

        # 检查文件是否存在
        if not os.path.exists(csv_path):
            print(f"警告：CSV文件不存在: {csv_path}")
            return {}

        data = {}
        encodings = ['gbk', 'gb2312', 'utf-8-sig', 'utf-8']

        for encoding in encodings:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    next(reader)  # 跳过表头
                    for row in reader:
                        if len(row) >= 2 and row[0].strip():
                            key = row[0].strip()
                            value = row[1].strip() if len(row) > 1 else ''
                            data[key] = value

                print(f"成功从CSV读取 {len(data)} 个字段 (编码: {encoding})")

                # 计算scope_3_emissions总和（如果CSV中没有）
                if 'scope_3_emissions' not in data:
                    scope3_total = 0
                    for i in range(1, 16):
                        key = f'scope_3_category_{i}_emissions'
                        if key in data and data[key]:
                            try:
                                scope3_total += float(data[key])
                            except ValueError:
                                pass
                    data['scope_3_emissions'] = str(round(scope3_total, 6))
                    print(f"计算得出 scope_3_emissions: {data['scope_3_emissions']}")

                return data

            except (UnicodeDecodeError, Exception) as e:
                continue

        print(f"错误：无法使用任何编码读取CSV文件")
        return {}

    def _parse_csv_sections(self, csv_path='减排行动统计.csv'):
        """
        按区域解析CSV文件，保留行号顺序
        用于提取表格数据（范围一、范围二三的排放源）

        Returns:
            dict: {'scope1_items': [...], 'scope2_3_items': [...]}
        """
        import os
        import re

        # 检查文件是否存在
        if not os.path.exists(csv_path):
            return {'scope1_items': [], 'scope2_3_items': []}

        encodings = ['gbk', 'gb2312', 'utf-8-sig', 'utf-8']

        for encoding in encodings:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = list(reader)

                # 查找区域标记的位置（不写死行号）
                scope1_start = None
                scope2_3_start = None

                for i, row in enumerate(rows):
                    if len(row) >= 1 and row[0]:
                        row_text = str(row[0])
                        # 范围一：匹配"范围一"+"直接"+"排放源"
                        if '范围一' in row_text and '直接' in row_text and '排放源' in row_text:
                            scope1_start = i
                            print(f"找到范围一标记在第 {i+1} 行: {row_text}")
                        # 范围二三：匹配"范围二"或"范围二三"+"排放源"（间接可能有编码问题，用更宽松的匹配）
                        elif ('范围二' in row_text or '范围二三' in row_text) and '排放源' in row_text:
                            # 确保不是范围一
                            if '范围一' not in row_text:
                                scope2_3_start = i
                                print(f"找到范围二三标记在第 {i+1} 行: {row_text}")

                # 解析范围一数据（从"范围一直接排放源"到"范围二三间接排放源"之前）
                scope1_items = []
                if scope1_start is not None:
                    end = scope2_3_start if scope2_3_start else len(rows)

                    # CSV数据结构：第1列=类别，第2列=排放源，第3列=设施
                    # template.docx期望：name=类别, emission=排放源, note=设施
                    for i in range(scope1_start + 2, end):  # +2 跳过标记行和表头行
                        if len(rows[i]) >= 2 and rows[i][0] and rows[i][1]:
                            category = str(rows[i][0]).strip()
                            source = str(rows[i][1]).strip()
                            facility = str(rows[i][2]).strip() if len(rows[i]) >= 3 else ''

                            # 跳过空值和标题行
                            if not source or source == '排放源' or source == category:
                                continue

                            scope1_items.append({
                                'name': category,
                                'emission': source,  # 排放源放在emission字段
                                'note': facility  # 设施放在note字段
                            })

                    print(f"解析范围一数据: {len(scope1_items)} 条记录")

                # 解析范围二三数据（从"范围二三间接排放源"开始）
                scope2_3_items = []
                if scope2_3_start is not None:
                    # CSV数据结构：第1列=类别，第2列=排放源，第3列=设施
                    # template.docx期望：name=类别, emission=排放源, note=设施
                    for i in range(scope2_3_start + 2, len(rows)):  # +2 跳过标记行和表头行
                        if len(rows[i]) >= 2 and rows[i][0] and rows[i][1]:
                            category = str(rows[i][0]).strip()
                            source = str(rows[i][1]).strip()
                            facility = str(rows[i][2]).strip() if len(rows[i]) >= 3 else ''

                            # 跳过空值和标题行
                            if not source or source == '排放源' or source == category:
                                continue

                            # 过滤掉范围一的重复数据（类别名包含"范围一"）
                            if '范围一' in category:
                                continue

                            scope2_3_items.append({
                                'name': category,
                                'emission': source,  # 排放源放在emission字段
                                'note': facility  # 设施放在note字段
                            })

                    print(f"解析范围二三数据: {len(scope2_3_items)} 条记录")

                return {
                    'scope1_items': scope1_items,
                    'scope2_3_items': scope2_3_items
                }

            except (UnicodeDecodeError, Exception) as e:
                print(f"解析CSV区域时出错 (编码 {encoding}): {e}")
                continue

        return {'scope1_items': [], 'scope2_3_items': []}

    def find_multiple_values_by_pattern(self, sheet_name, patterns, search_direction='right',
                                      max_distance=3, require_numeric=False):
        """
        根据模式匹配查找多个值，用于处理复杂的表格结构

        Args:
            sheet_name: 工作表名称
            patterns: 模式列表，支持正则表达式
            search_direction: 搜索方向
            max_distance: 搜索最大距离（单元格数）
            require_numeric: 是否要求数值结果

        Returns:
            找到的值列表
        """
        import re
        results = []

        if not self.workbook or self.file_type != 'excel':
            return results

        try:
            sheet = self.workbook[sheet_name]
            if sheet_name not in self.workbook.sheetnames:
                return results

            for pattern in patterns:
                # 编译正则表达式
                try:
                    regex = re.compile(pattern, re.IGNORECASE)
                except re.error as e:
                    print(f"正则表达式错误 '{pattern}': {e}")
                    continue

                # 搜索匹配的单元格
                matched_cells = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and regex.search(str(cell.value)):
                            matched_cells.append(cell)

                # 为每个匹配的单元格查找相邻值
                for cell in matched_cells:
                    for distance in range(1, max_distance + 1):
                        value_cell = None

                        if search_direction == 'right':
                            if cell.column + distance <= sheet.max_column:
                                value_cell = sheet.cell(row=cell.row, column=cell.column + distance)
                        elif search_direction == 'below':
                            if cell.row + distance <= sheet.max_row:
                                value_cell = sheet.cell(row=cell.row + distance, column=cell.column)

                        if value_cell and value_cell.value is not None:
                            # 检查是否需要数值
                            if require_numeric and not isinstance(value_cell.value, (int, float)):
                                # 尝试转换为数字
                                try:
                                    numeric_value = float(str(value_cell.value))
                                    results.append(numeric_value)
                                except (ValueError, TypeError):
                                    continue
                            else:
                                results.append(value_cell.value)
                            break

        except Exception as e:
            print(f"模式匹配查找时出错: {e}")

        return results

    def get_table_data_by_labels(self, sheet_name, row_labels, column_labels,
                                header_row=None, data_start_row=None):
        """
        根据行标签和列标签提取表格数据

        Args:
            sheet_name: 工作表名称
            row_labels: 行标签列表
            column_labels: 列标签列表
            header_row: 表头行位置
            data_start_row: 数据开始行

        Returns:
            字典格式的表格数据
        """
        if not self.workbook or self.file_type != 'excel':
            return {}

        try:
            sheet = self.workbook[sheet_name]
            if sheet_name not in self.workbook.sheetnames:
                return {}

            result = {}

            # 查找行标签位置
            row_positions = {}
            for label in row_labels:
                for row in range(1, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value and label.lower() in str(cell.value).lower():
                            row_positions[label] = col
                            break
                    if label in row_positions:
                        break

            # 查找列标签位置
            if header_row:
                col_positions = {}
                for label in column_labels:
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=header_row, column=col)
                        if cell.value and label.lower() in str(cell.value).lower():
                            col_positions[label] = col
                            break
            else:
                # 如果没有指定表头行，搜索整个工作表
                col_positions = {}
                for label in column_labels:
                    for row in range(1, sheet.max_row + 1):
                        for col in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row, column=col)
                            if cell.value and label.lower() in str(cell.value).lower():
                                col_positions[label] = col
                                break
                        if label in col_positions:
                            break

            # 提取数据
            data_start = data_start_row or 2
            for row_label, row_col in row_positions.items():
                result[row_label] = {}
                for col_label, col_num in col_positions.items():
                    # 从数据开始行向下查找
                    for row in range(data_start, sheet.max_row + 1):
                        if sheet.cell(row=row, column=row_col).value:  # 找到有数据的行
                            cell = sheet.cell(row=row, column=col_num)
                            result[row_label][col_label] = cell.value
                            break
                    else:
                        result[row_label][col_label] = None

            return result

        except Exception as e:
            print(f"提取表格数据时出错: {e}")
            return {}

    def _parse_csv_table_section(self, csv_data, section_label, emission_type='scope1'):
        """
        解析CSV中特定区域的表格数据

        Args:
            csv_data: 原始CSV数据字典 {参数: 值}
            section_label: 区域标签（如"范围一直接排放源"）
            emission_type: 排放类型 ('scope1', 'scope2_3')

        Returns:
            包含name, emission, note的字典列表
        """
        import re
        items = []

        # 查找区域起始位置
        section_start = None
        for key in csv_data.keys():
            if section_label in str(key):
                section_start = key
                break

        if not section_start:
            return items

        # 解析该区域下的数据
        # 根据CSV结构，数据按照 "类别-排放源-具体排放源" 的层级组织
        current_category = None
        emission_sources = []

        # 范围一的类别映射
        scope1_categories = {
            '固定燃烧': '固定燃料燃烧',
            '移动燃烧': '移动燃料燃烧',
            '散逸排放': '散逸排放',
            '过程排放': '工艺过程排放'
        }

        # 范围二、三的类别映射
        scope2_3_categories = {
            '外购热力': '外购热力产生的排放',
            '时间序列': '外购热力产生的排放',
        }

        # 收集所有相关的排放源数据
        for key, value in csv_data.items():
            if not value or value.strip() == '':
                continue

            # 根据排放类型选择相应的类别
            if emission_type == 'scope1':
                for category, display_name in scope1_categories.items():
                    if category in str(key):
                        items.append({
                            'name': display_name,
                            'emission': '0.00',
                            'note': value
                        })
                        break
            else:  # scope2_3
                # 范围二、三的处理
                if '外购电力' in str(key) or '外购热力' in str(key):
                    items.append({
                        'name': value,  # 使用CSV中的值作为排放源名称
                        'emission': '0.00',
                        'note': '外购能源'
                    })
                elif '上游' in str(key) or '产生的排放' in str(key):
                    # 上游排放源
                    items.append({
                        'name': value,
                        'emission': '0.00',
                        'note': '上游排放'
                    })
                elif any(x in str(key) for x in [' purchased_goods', 'category_1', 'category_2']):
                    items.append({
                        'name': value,
                        'emission': '0.00',
                        'note': '商品和服务'
                    })

        return items

    def extract_data(self):
        """
        提取所有模板需要的数据，返回包含32个变量的字典。
        优先从CSV文件读取，如果CSV不存在则从Excel文件提取。
        """
        # 默认值字典 - 用于数据源中不存在的情况
        default_values = {
            'company_profile': '待补充公司简介信息',
            'legal_person': '待补充',
            'registered_address': '待补充注册地址',
            'date_of_establishment': '待补充',
            'registered_capital': '待补充',
            'Unified_Social_Credit_Identifier': '待补充',
            'deadline': '待补充',
            'evaluation_level': '待评估',
            'evaluation_score': '待评估',
            'scope_of_business': '待补充经营范围',
            'source_file': self.filepath if hasattr(self, 'filepath') else '未知',
            'GWP_Value_Reference_Document': '2021年IPCC第六次评估报告AR6',
            'rule_file': '企业温室气体排放核算与报告指南',
        }

        # 初始化数据字典
        data = {}
        data.update(default_values)

        # ========== 优先尝试从CSV文件读取所有数据 ==========
        import os
        csv_path = '减排行动统计.csv'
        if os.path.exists(csv_path):
            csv_data = self.read_emission_data_csv(csv_path)
            if csv_data:
                data.update(csv_data)
                print(f"从CSV文件成功读取 {len(csv_data)} 个变量")

                # ========== 格式化数字：保留两位小数，添加千分位分隔符 ==========
                def format_number(value):
                    """格式化数字：保留两位小数，添加千分位分隔符"""
                    try:
                        return f"{float(value):,.2f}"
                    except (ValueError, TypeError):
                        return "0.00"

                # 格式化所有排放数据（原始键名，用于模板渲染）
                emission_keys = [
                    'scope_1_emissions',
                    'scope_2_location_based_emissions',
                    'scope_2_market_based_emissions',
                    'scope_3_emissions',
                    'scope_3_category_1_emissions',
                    'scope_3_category_2_emissions',
                    'scope_3_category_3_emissions',
                    'scope_3_category_4_emissions',
                    'scope_3_category_5_emissions',
                    'scope_3_category_6_emissions',
                    'scope_3_category_7_emissions',
                    'scope_3_category_9_emissions',
                    'scope_3_category_10_emissions',
                    'scope_3_category_12_emissions',
                ]
                for key in emission_keys:
                    if key in data:
                        data[key] = format_number(data[key])

                # ========== 构建表格数据列表（使用按区域解析的方法）==========
                section_data = self._parse_csv_sections(csv_path)

                # 获取范围一和范围二三的表格数据
                scope1_items = section_data.get('scope1_items', [])
                scope2_3_items_raw = section_data.get('scope2_3_items', [])

                # 构建最终的scope2_3_items列表
                # 1. 首先添加范围二的总量数据（如果有）
                scope2_3_items = []
                scope2_location = data.get('scope_2_location_based_emissions', '0.00')
                scope2_market = data.get('scope_2_market_based_emissions', '0.00')

                if scope2_location != '0.00':
                    scope2_3_items.append({
                        'name': '范围二：能源间接温室气体排放（基于位置）',
                        'emission': scope2_location,
                        'note': '外购电力和热力'
                    })
                if scope2_market != '0.00':
                    scope2_3_items.append({
                        'name': '范围二：能源间接温室气体排放（基于市场）',
                        'emission': scope2_market,
                        'note': '外购电力和热力'
                    })

                # 2. 添加范围三分类数据（如果有）
                scope3_total_items = [
                    ('外购商品和服务的上游产生的排放', 'scope_3_category_1_emissions', '原材料采购'),
                    ('资本货物产生的排放', 'scope_3_category_2_emissions', '设备设施建设'),
                    ('燃料和能源相关逸出排放', 'scope_3_category_3_emissions', '外购电力热力上游排放'),
                    ('上下游运输和配送产生的排放', 'scope_3_category_4_emissions', '物流运输'),
                    ('运营中产生的废弃物产生的排放', 'scope_3_category_5_emissions', '废弃物处理'),
                    ('员工商务差旅产生的排放', 'scope_3_category_6_emissions', '商务出行'),
                    ('员工上下班通勤产生的排放', 'scope_3_category_7_emissions', '员工通勤'),
                    ('运营中输入的运输和配送产生的排放', 'scope_3_category_9_emissions', '原材料和产品运输'),
                    ('已售产品的使用过程产生的排放', 'scope_3_category_10_emissions', '产品使用阶段'),
                    ('已售产品的报废处理产生的排放', 'scope_3_category_12_emissions', '产品回收处理'),
                ]

                for name, emission_key, note in scope3_total_items:
                    emission_value = data.get(emission_key, '0.00')
                    if emission_value != '0.00':
                        scope2_3_items.append({
                            'name': name,
                            'emission': emission_value,
                            'note': note
                        })

                # 3. 最后添加从CSV解析的详细排放源数据
                scope2_3_items.extend(scope2_3_items_raw)

                data['scope1_items'] = scope1_items
                data['scope2_3_items'] = scope2_3_items

                # 为了向后兼容，保留 items 列表（使用范围二三数据）
                data['items'] = scope2_3_items

                # ========== 键名映射：为AIService添加别名（用于AI摘要生成）==========
                # CSV使用的键名 -> AIService期望的键名
                key_mapping = {
                    'scope_1_emissions': 'scope_1',
                    'scope_2_location_based_emissions': 'scope_2_location',
                    'scope_2_market_based_emissions': 'scope_2_market',
                    'scope_3_emissions': 'scope_3',
                }

                for csv_key, ai_key in key_mapping.items():
                    if csv_key in data:
                        data[ai_key] = data[csv_key]

                # 计算总排放量（用于AI摘要）
                try:
                    s1 = float(data.get('scope_1_emissions', 0))
                    s2_loc = float(data.get('scope_2_location_based_emissions', 0))
                    s3 = float(data.get('scope_3_emissions', 0))
                    total_loc = s1 + s2_loc + s3

                    s2_mkt = float(data.get('scope_2_market_based_emissions', 0))
                    total_mkt = s1 + s2_mkt + s3

                    data['total_emission_location'] = format_number(total_loc)
                    data['total_emission_market'] = format_number(total_mkt)

                    # 提取年份
                    period = data.get('reporting_period', '')
                    import re
                    year_match = re.search(r'(\d{4})', str(period))
                    data['report_year'] = year_match.group(1) if year_match else '2024'
                except (ValueError, TypeError):
                    data['total_emission_location'] = "0.00"
                    data['total_emission_market'] = "0.00"
                    data['report_year'] = '2024'

                return data

        # ========== 如果CSV不存在，使用Excel数据（向后兼容） ==========
        # 处理CSV文件（减排行动统计数据）
        if self.file_type == 'csv' and ('减排行动' in str(self.filepath) or 'GHG' in str(self.filepath)):
            emission_reductions = self.read_to_list_of_dicts(skip_empty_rows=True)
            data['emission_reductions'] = emission_reductions
            data['file_type'] = 'csv'
            print(f"从CSV文件提取减排行动数据，共 {len(emission_reductions)} 条记录")
            return data

        # 处理Excel文件（温室气体排放数据）
        if not self.workbook or self.file_type != 'excel':
            return data

        # 尝试多个可能的工作表名称
        main_sheet_candidates = ['温室气体盘查清册', '温室气体盘查清册 (2)']
        main_sheet = None
        for candidate in main_sheet_candidates:
            if candidate in self.workbook.sheetnames:
                main_sheet = candidate
                break

        if not main_sheet:
            print("警告：未找到主要工作表")
            return data

        table_sheet = '表1温室气体盘查表'
        
        # 使用新的find_value_by_label方法替代硬坐标定位
        # 从主要工作表中提取元数据
        company_name = self.find_value_by_label(main_sheet, '组织名称：') 
        report_period = self.find_value_by_label(main_sheet, '盘查覆盖周期:') 
        # 从报告周期中提取年份（假设格式为"2024年1月1日至2024年12月31日"）
        report_year = '2024'  # 直接提取年份

        # 获取范围一排放量
        scope_1 = None
        try:
            sheet = self.workbook[table_sheet]
            # 遍历表格找到'总排放量'所在行
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == '总排放量':
                        # 根据用户反馈，总排放量这一行的数据与正上方单元格一一对应
                        # 因此我们需要获取当前行各列的值，然后将这些值与上一行的标签对应
                        # 这里我们主要关注范围一对应的排放量
                        current_row = cell.row
                        # 假设范围一的标签在B列（根据之前的调试发现）
                        # 检查上一行B列是否包含'范围一'
                        prev_row_cell_b = sheet.cell(row=current_row-1, column=2)
                        if prev_row_cell_b.value and '范围一' in str(prev_row_cell_b.value):
                            # 获取当前行B列的值作为scope_1
                            scope_1 = sheet.cell(row=current_row, column=2).value
                            print(f"从表1温室气体盘查表获取scope_1值(总排放量行上方对应范围一): {scope_1}")
                            break
                if scope_1 is not None:
                    break
            
            # 如果没找到，回退到使用find_value_by_label方法
            if scope_1 is None:
                scope_1 = self.find_value_by_label(table_sheet, '总排放量')
                print(f"回退到查找总排放量右侧值作为scope_1: {scope_1}")
        except Exception as e:
            print(f"获取scope_1值时出错: {e}")
        
        # 提取范围二排放量
        # 使用find_value_by_label方法替代硬坐标
        scope_2_location = self.find_value_by_label(table_sheet, '基于位置')
        scope_2_market = self.find_value_by_label(table_sheet, '基于市场')
        
        # 如果直接查找失败，回退到原始方法
        if scope_2_location is None:
            scope_2_location = self.find_value_by_label(table_sheet, '范围二')
        
        # 提取范围三排放量
        scope_3 = None
        try:
            # 优先使用find_value_by_label方法
            scope_3 = self.find_value_by_label(table_sheet, '范围三')
            
            if scope_3 is None:
                # 如果直接查找失败，尝试查找包含"范围三"的单元格
                sheet = self.workbook[table_sheet]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and '范围三' in str(cell.value):
                            # 检查右侧和下方的单元格
                            right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                            below_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                            
                            # 优先尝试右侧单元格（如果是数值或总量）
                            if right_cell.value is not None:
                                if isinstance(right_cell.value, (int, float)):
                                    scope_3 = right_cell.value
                                elif right_cell.value == '总量':
                                    # 如果右侧是总量，获取总量下方的值
                                    total_below = sheet.cell(row=right_cell.row + 1, column=right_cell.column)
                                    if total_below.value is not None:
                                        scope_3 = total_below.value
                            # 如果右侧没有找到，尝试下方单元格
                            elif below_cell.value is not None and isinstance(below_cell.value, (int, float)):
                                scope_3 = below_cell.value
                            break
                    if scope_3 is not None:
                        break
        except Exception as e:
            print(f"查找范围三数据时出错: {e}") 
        
        # 提取总排放量（基于位置）和总排放量（基于市场）
        # 使用find_value_by_label方法替代硬坐标
        total_emission_location = None
        total_emission_market = None
        try:
            # 首先计算预期的总排放量范围，用于验证找到的值是否合理
            expected_total_location = None
            expected_total_market = None
            if scope_1 is not None and scope_2_location is not None and scope_3 is not None:
                expected_total_location = float(scope_1) + float(scope_2_location) + float(scope_3)
                expected_total_market = float(scope_1) + float(scope_2_market) + float(scope_3)
                print(f"预期总排放量范围: 位置={expected_total_location}, 市场={expected_total_market}")
            
            # 使用find_value_by_label方法查找总排放量
            total_emission_location = self.find_value_by_label(table_sheet, '总排放量')
            
            # 尝试查找包含'基于位置'和'总量'的区域
            if total_emission_location is None:
                sheet = self.workbook[table_sheet]
                # 记录所有可能的候选值
                potential_totals = []
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_text = str(cell.value)
                            # 查找包含'总量'的单元格
                            if '总量' in cell_text:
                                # 记录这个位置附近的所有数值
                                for r_offset in range(-4, 5):
                                    for c_offset in range(-4, 5):
                                        check_row = cell.row + r_offset
                                        check_col = cell.column + c_offset
                                        if 1 <= check_row <= sheet.max_row and 1 <= check_col <= sheet.max_column:
                                            check_cell = sheet.cell(row=check_row, column=check_col)
                                            if check_cell.value is not None and isinstance(check_cell.value, (int, float)):
                                                # 检查这个数值是否接近预期的总排放量
                                                is_large_value = check_cell.value > 1000000  # 假设总排放量大于100万
                                                
                                                # 检查是否有'基于位置'或'基于市场'的标注
                                                has_location = False
                                                has_market = False
                                                for context_r in range(max(1, check_row-3), min(sheet.max_row, check_row+4)):
                                                    for context_c in range(max(1, check_col-5), min(sheet.max_column, check_col+6)):
                                                        context_cell = sheet.cell(row=context_r, column=context_c)
                                                        if context_cell.value is not None:
                                                            context_text = str(context_cell.value)
                                                            if '基于位置' in context_text:
                                                                has_location = True
                                                            elif '基于市场' in context_text:
                                                                has_market = True
                                                
                                                # 添加到候选列表
                                                if is_large_value and (has_location or has_market):
                                                    potential_totals.append({
                                                        'value': check_cell.value,
                                                        'row': check_row,
                                                        'col': check_col,
                                                        'is_location': has_location,
                                                        'is_market': has_market
                                                    })
                
                # 从候选值中选择最接近预期值的
                if potential_totals:
                    # 按与预期值的接近程度排序
                    if expected_total_location is not None:
                        potential_totals.sort(key=lambda x: abs(x['value'] - expected_total_location) if x['is_location'] else float('inf'))
                    # 选择第一个合适的基于位置的候选值
                    for candidate in potential_totals:
                        if candidate['is_location']:
                            total_emission_location = candidate['value']
                            print(f"从候选值中选择总排放量（基于位置）在第{candidate['row']}行第{candidate['col']}列: {candidate['value']}")
                            break
                    
                    # 选择第一个合适的基于市场的候选值
                    for candidate in potential_totals:
                        if candidate['is_market']:
                            total_emission_market = candidate['value']
                            print(f"从候选值中选择总排放量（基于市场）在第{candidate['row']}行第{candidate['col']}列: {candidate['value']}")
                            break
            
            # 方法3: 如果仍然找不到，直接使用计算值
            if total_emission_location is None and expected_total_location is not None:
                total_emission_location = expected_total_location
                print(f"使用计算值作为总排放量（基于位置）: {total_emission_location}")
            
            if total_emission_market is None and expected_total_market is not None:
                total_emission_market = expected_total_market
                print(f"使用计算值作为总排放量（基于市场）: {total_emission_market}")
        except Exception as e:
            print(f"获取总排放量时出错: {e}")
        
        # 将所有数据打包成一个标准字典 
        data = { 
            'company_name': company_name, 
            'report_year': report_year, 
            'scope_1': scope_1,  # 范围一排放量
            'scope_2_location': scope_2_location,  # 范围二排放量（基于位置）
            'scope_2_market': scope_2_market,      # 范围二排放量（基于市场）
            'scope_3': scope_3,                     # 范围三排放量
            'total_emission_location': total_emission_location,  # 总排放量（基于位置）
            'total_emission_market': total_emission_market,        # 总排放量（基于市场）
            'file_type': 'excel'
        } 
        
        print(f"数据提取完成: {data}") 
        return data
        
    def extract_all_data(self):
        """ 
        提取所有数据，包括从Excel文件和CSV文件中提取的内容。
        整合了温室气体排放数据和减排行动数据。
        """
        # 默认返回值
        result = {
            'greenhouse_gas_data': {},
            'emission_reductions': []
        }
        
        # 首先处理Excel文件（如果有）
        if self.file_type == 'excel' and self.workbook:
            result['greenhouse_gas_data'] = self.extract_data()
        
        # 检查是否有减排行动CSV文件
        import os
        csv_file_path = 'D:\\my_report_generator\\减排行动统计.csv'
        if os.path.exists(csv_file_path):
            # 创建一个临时的ExcelDataReader实例来读取CSV文件
            csv_reader = ExcelDataReader(csv_file_path)
            result['emission_reductions'] = csv_reader.read_to_list_of_dicts(skip_empty_rows=True)
            print(f"成功从CSV文件读取 {len(result['emission_reductions'])} 条减排行动数据")
        
        return result 

# 测试函数
if __name__ == "__main__": 
    # 测试1: Excel数据读取
    reader = ExcelDataReader('test_data.xlsx') 
    data = reader.extract_data() 
    print("--- 测试 data_reader.py ---\n", data)
    
    # 测试2: find_value_by_label方法
    if reader.workbook:
        test_value = reader.find_value_by_label('温室气体盘查清册', '组织名称：')
        print(f"\n--- 测试 find_value_by_label 方法 ---\n组织名称: {test_value}")
    
    # 测试3: 读取CSV文件（如果存在）
    csv_file = '减排行动统计.csv'
    if os.path.exists(csv_file):
        csv_reader = ExcelDataReader(csv_file)
        csv_data = csv_reader.read_to_list_of_dicts()
        print(f"\n--- 测试 CSV 读取功能 ---\n读取到 {len(csv_data)} 行数据")
        if csv_data:
            print("数据示例:", csv_data[:2])  # 打印前两行数据
    
    # 测试4: 读取Excel工作表为列表字典
    if reader.workbook:
        excel_list_data = reader.read_to_list_of_dicts('表1温室气体盘查表')
        print(f"\n--- 测试 Excel 列表字典读取功能 ---\n读取到 {len(excel_list_data)} 行数据")
        if excel_list_data:
            print("数据示例:", excel_list_data[:2])  # 打印前两行数据